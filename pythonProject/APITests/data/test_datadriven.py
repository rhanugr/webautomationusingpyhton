import requests
import json
import jsonpath
import openpyxl

def test_data():
    api_url = "http://www.thetestingworldapi.com/api/studentsDetails"
    file = open('C:/Hanamanta_Data/Datadriven/test.json')
    json_request = json.loads(file.read())
    # response = requests.post(api_url,json_request)
    # print(response.status_code)
    # print(response.text)
    book = openpyxl.load_workbook('C:\\Hanamanta_Data\\testdata.xlsx')
    sheet = book.active
    row = sheet.max_row
    for i in range(2, row+1):
        cell_first_name = sheet.cell(row=i , column=1)
        cell_middle_name = sheet.cell(row=i, column=2)
        cell_last_name = sheet.cell(row=i, column=3)
        cell_date_of_birth = sheet.cell(row=i, column=4)

        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_date_of_birth.value

        response = requests.post(api_url,json_request)
        print('Status Code : ',response.status_code)
        print(response.text)




import openpyxl


book = openpyxl.load_workbook('C:\\Hanamanta_Data\\testdata.xlsx')
sheet = book.active
row =   sheet.max_row

for i in range(2,row+1):
    cell_first_name = sheet.cell(row = i,column=1)
    cell_middle_name = sheet.cell(row=i, column=2)
    cell_last_name = sheet.cell(row=i, column=3)
    cell_date_of_birth = sheet.cell(row=i, column=4)
    print(cell_first_name.value)
    print(cell_last_name.value)
    print(cell_middle_name.value)
    print(cell_date_of_birth.value)








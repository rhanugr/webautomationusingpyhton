import openpyxl

book = openpyxl.open("C:\\Hanamanta_Data\\test.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column= 2)
print(cell.value)

setval = sheet.cell(row=2, column= 2).value = "Prateek"
print(setval)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['B5'].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value =="T3":
        for j in range(1, sheet.max_column+1):
            print(sheet.cell(row=i , column=j).value)
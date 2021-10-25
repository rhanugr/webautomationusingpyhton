import json

import openpyxl
import requests


def test_student_details():
    #1 api_url should be available

    api_url = 'http://www.thetestingworldapi.com/api/studentsDetails'

    #2.Should have API body which have to sent to server Get it from Json (using notpad++) edit and save

# ''' {
#   "first_name": "sample string 2",
#   "middle_name": "sample string 3",
#   "last_name": "sample string 4",
#   "date_of_birth": "sample string 5"
#     }
# '''
#     #3.open the json file store it in file

    file = open('C:/Hanamanta_Data/Datadriven/student.json')

    #Loads it into Json format
    json_request = json.loads(file.read())

    #Post request use to create a data
    # store_response = requests.post(api_url ,json_request)
    #
    # print(store_response.status_code)
    # print(store_response.text)
    #Load the workbook for read the data from excel
    book = openpyxl.load_workbook('C:/Hanamanta_Data/testdata.xlsx')
    #Go to the Active sheet
    sheet = book.active
    row = sheet.max_row
    for i in range(2,row+1):
        cell_first = sheet.cell(row=i, column=1)
        cell_midle = sheet.cell(row=i, column=2)
        cell_lastname = sheet.cell(row=i , column=3)
        cell_dob = sheet.cell(row = i, column=4)
        json_request['first_name'] = cell_first.value
        json_request['middle_name'] = cell_midle.value
        json_request['last_name'] = cell_lastname.value
        json_request['date_of_birth'] = cell_dob.value

        store_response = requests.post(api_url, json_request)
        print(store_response.status_code)
        print(store_response.text)



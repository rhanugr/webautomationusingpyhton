import openpyxl


class Lib():

    def __init__(self, FilePath , SheetName):
        global book
        global sheet
        book = openpyxl.load_workbook(FilePath)
        sheet =  book.active


    def fetch_row_count(self):
        row = sheet.max_row
        return row

    def fetch_column_count(self):
        cols = sheet.max_column
        return cols

    def fecth_key_name(self):
        c =sheet.max_column
        li =[]
        for i in range(1,c+1):
            cell = sheet.cell(row=1,column=i)
            li.insert(i-1,cell.value)
            return li

    def update_requets_with_data(self,rowNumber,jsonRequest,keyList):
        c = sheet.max_column
        for i in range(1,c+1):
            cell = sheet.cell(row=rowNumber,column=i)
            jsonRequest[keyList[i-1]]=cell.value
            return jsonRequest






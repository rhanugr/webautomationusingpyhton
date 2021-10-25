import openpyxl

class Test:
    def fecth_key_name(self):
        book = openpyxl.load_workbook('C:/Hanamanta_Data/testdata.xlsx')
        sheet = book.active

        c = sheet.max_column
        li = []
        for i in range(1, c + 1):
            cell = sheet.cell(row=1, column=i)
            li.insert(i - 1, cell.value)
            print(li)

    # print(li)
obj =  Test()
obj.fecth_key_name()

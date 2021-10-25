import openpyxl

workbook = openpyxl.load_workbook("C:/Hanamanta_Data/testdata.xlsx")

sheet = workbook.active
rw =  sheet.max_row
col = sheet.max_column
for i in range(2,rw+1):
    for j in range(2,col+1):
        cell_value = sheet.cell(row= i, column=j)
        print(cell_value.value)


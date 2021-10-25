import openpyxl


class Test:

    def update_requets_with_data(self):
        book = openpyxl.load_workbook('C:/Hanamanta_Data/testdata.xlsx')
        sheet = book.active
        c = sheet.max_column
        for i in range(1, c + 1):
            cell = sheet.cell(row=1, column=i)
            jsonRequest[keyList[i - 1]] = cell.value
            return jsonRequest

